"""
This is CACC traffic flow optimization code with distributed Q-learning.
It consists of 2 major part, which are controllers applied in CARLA and the learning algorithms.
Authored by Chuheng Wu
"""


import threading
import glob
import os
import sys
import numpy as np
import random

try:
    sys.path.append(glob.glob('../carla/dist/carla-*%d.%d-%s.egg' % (
        sys.version_info.major,
        sys.version_info.minor,
        'win-amd64' if os.name == 'nt' else 'linux-x86_64'))[0])
except IndexError:
    pass

import carla
import time
import math

from openpyxl import Workbook
from openpyxl import load_workbook


wb = Workbook()
wb = load_workbook(r"C:\Users\Articial_Idiot\Desktop\data.xlsx")
ws = wb.active
note = ws

class CACC(threading.Thread):
    def __init__(self, note):
        threading.Thread.__init__(self)
        self.wb = Workbook()
        # self.wb = load_workbook(r"C:\Users\Articial_Idiot\Desktop\data.xlsx")
        self.ws = self.wb.active
        self.wb = note
        self.note = wb.active
        self.actor_list = []
        self.lane1_list = []
        self.lane2_list = []
        self.lane3_list = []
        self.lane4_list = []
        self.thread_list = []
        self.ave_velocity = 0
        self.destroyed_num = 0
        self.shock = 0
        self.frame = 0
        self._flag = 0
        self.crash = 0
        self.cond = threading.Condition()

        self.episode = 0
        self.reward_sum = 0
        self.call = 0
        self.v_init = 0
        self.q_table = np.random.uniform(low=-1, high=1, size=(4 ** 5, 4))

        self.world = 0
        self.client = 0


    def spawn_vehicle(self, lane):
        sp1 = carla.Transform(carla.Location(x=335, y=9.5, z=0.7), carla.Rotation(yaw=180))
        sp2 = carla.Transform(carla.Location(x=335, y=14, z=0.7), carla.Rotation(yaw=180))
        sp3 = carla.Transform(carla.Location(x=335, y=17.6, z=0.7), carla.Rotation(yaw=180))
        sp4 = carla.Transform(carla.Location(x=335, y=20.2, z=0.7), carla.Rotation(yaw=180))
        bp = self.world.get_blueprint_library().find('vehicle.tesla.model3')
        bp.set_attribute('color', str(random.randint(0,255))+','+str(random.randint(0,255))+','+str(random.randint(0,255)))
        try:
            if lane == 1:
                a = self.world.spawn_actor(bp, sp1)
            if lane == 2:
                a = self.world.spawn_actor(bp, sp2)
            if lane == 3:
                a = self.world.spawn_actor(bp, sp3)
            if lane == 4:
                a = self.world.spawn_actor(bp, sp4)
            Lock = threading.Lock()
            Lock.acquire()
            self.actor_list.append(a)
            ID = len(self.actor_list) - 1
            # print("vehicle " + str(ID) + " has been spawned")
            Lock.release()
        except:
            pass

    def crash_detection(self):
        while True:
            if len(self.actor_list)>0:
                try:
                    Lock  = threading.Lock()
                    Lock.acquire()
                    for i in range(1, len(self.actor_list)):
                        for j in range(1, len(self.actor_list)):
                            if not(abs(self.actor_list[i].get_transform().x - self.actor_list[j].get_transform().x) < 0.5 and abs(self.actor_list[i].get_transform().y - self.actor_list[j].get_transform().y) < 0.3):
                                if abs(self.actor_list[i].get_transform().x - self.actor_list[j].get_transform().x) < 4.9 and abs(self.actor_list[i].get_transform().y - self.actor_list[j].get_transform().y) < 1.9:
                                    self.crash = 1
                                    print("crashed!")
                    Lock.release()
                except:
                    pass
            time.sleep(1)

    def x_controller(self, a, a_1, eba, ebx, ai_1_list_a, qi_buffer):
        kp = 0.13
        kd1 = 0.8
        kd2 = 0
        ki = 0
        k = 1
        k_p = 1.08
        k_i = 1.29
        k_d = 0.00
        k_pp = 0.05
        k_ii = 4
        k_dd = 0
        delay_frames = 3
        stand_still_distance = 5
        h = 2
        T = 0
        B = 0
        qi = 0
        adi = 0
        ebx.append(
            -a_1.get_transform().location.x + a.get_transform().location.x - stand_still_distance + h * a.get_velocity().x)
        ai_1_list_a.append(-a_1.get_acceleration().x)

        if len(ebx) > 2 and len(ai_1_list_a) > delay_frames + 1:
            dex = (ebx[-1] - ebx[-2]) / 0.03
            ddex = (ebx[-1] + ebx[-3] - 2 * ebx[-2]) / 0.009
            iex = sum(ebx) * 0.03
            qi = k * ai_1_list_a[-delay_frames - 1] + kp * ebx[-1] + kd1 * dex + kd2 * ddex + iex * ki
            # ebx.pop(0)
            ai_1_list_a.pop(0)

        elif len(ebx) <= 2 or len(ai_1_list_a) <= delay_frames + 1:
            qi = 0
        qi_buffer.append(qi)
        if len(qi_buffer) > 7:
            adi = sum(qi_buffer)/8
            qi_buffer.pop(0)
        else:
            adi = sum(qi_buffer)/len(qi_buffer)



        error = adi + a.get_acceleration().x
        eba.append(error)
        if len(eba) >= 2:
            _de = (eba[-1] - eba[-2]) / 0.03
            _ie = sum(eba) * 0.03
        else:
            _de = 0.0
            _ie = 0.0
            error = 0
        T = np.clip((k_p * error) + (k_d * _de) + (k_i * _ie), 0, 1)
        if T < 0:
            T = 0
        if qi < 0 and T == 0:
            B = -np.clip((k_pp * error) + (k_dd * _de) + (k_ii * _ie), -1, 0)
        else:
            B = 0
        if -a.get_velocity().x >= 25 and adi > 0:
            T = 0
        if a.get_velocity().x < 0:
            h_ = (-a_1.get_transform().location.x + a.get_transform().location.x - stand_still_distance) / -a.get_velocity().x
        else:
            h_ = 15
        if h_ <= 0.5 and a.get_velocity().x < 1:
            T = 0
            B = 1
            self.shock += 1

        return T, B, eba, ebx, ai_1_list_a, qi_buffer

    def y_controller(self, actor, lane_t, ed_list):
        kp = 0.01
        ki = 0
        kd = 27
        ax = actor.get_transform().location.x
        ay = actor.get_transform().location.y

        if lane_t == 1:
            distance = ay - 0.00618 * ax - 7.43
        elif lane_t == 2:
            distance = ay - 0.00618 * ax - 11.03
        elif lane_t == 3:
            distance = ay - 0.00618 * ax - 14.63
        elif lane_t == 4:
            distance = ay - 0.00618 * ax - 18.23
        else:
            distance = 0

        ed_list.append(distance)
        if len(ed_list) > 1:
            steering = np.clip(kp * distance + kd * (ed_list[-1] - ed_list[-2]) * 0.03 + ki * sum(ed_list), -1, 1)
        else:
            steering = 0
        return steering, ed_list

    def rot(self, deg):
        if deg >= 0 and deg <= 180:
            r = math.radians(deg - 180)
        if deg < 0 and deg >= -180:
            r = math.radians(180 + deg)
        return r

    def lane_ob(self, actor):
        lane_r = 0
        lane_l = 0
        list_o = []
        list_l = []
        list_r = []
        actor_1 = 0
        actor_r = 0
        actor_l = 0
        actor_r_1 = 0
        actor_l_1 = 0
        ay = actor.get_transform().location.y
        ax = actor.get_transform().location.x


        distance = ay - 0.00618 * ax
        if abs(distance - 7.43) < 1.8:
            lane_o = 1
            lane_l = 2
            lane_r = 0
        elif abs(distance - 11.03) < 1.8:
            lane_o = 2
            lane_r = 1
            lane_l = 3
        elif abs(distance - 14.63) < 1.8:
            lane_o = 3
            lane_r = 2
            lane_l = 4
        elif abs(distance - 18.23) < 1.8:
            lane_o = 4
            lane_r = 3
            lane_l = 0
        else:
            lane_o = 0
            # print("lane_o can't be detected")

        """
        get the 5 vehicles
        """
        if lane_o == 1:
            for i in self.actor_list:
                if abs(i.get_transform().location.y - 0.00618 * i.get_transform().location.x - 7.43) < 1.8:
                    list_o.append(i)
                if abs(i.get_transform().location.y - 0.00618 * i.get_transform().location.x - 11.03) < 1.8:
                    list_l.append(i)
        if lane_o == 2:
            for i in self.actor_list:
                if abs(i.get_transform().location.y - 0.00618 * i.get_transform().location.x - 11.03) < 1.8:
                    list_o.append(i)
                if abs(i.get_transform().location.y - 0.00618 * i.get_transform().location.x - 14.63) < 1.8:
                    list_l.append(i)
                if abs(i.get_transform().location.y - 0.00618 * i.get_transform().location.x - 7.43) < 1.8:
                    list_r.append(i)
        if lane_o == 3:
            for i in self.actor_list:
                if abs(i.get_transform().location.y - 0.00618 * i.get_transform().location.x - 14.63) < 1.8:
                    list_o.append(i)
                if abs(i.get_transform().location.y - 0.00618 * i.get_transform().location.x - 18.23) < 1.8:
                    list_l.append(i)
                if abs(i.get_transform().location.y - 0.00618 * i.get_transform().location.x - 11.03) < 1.8:
                    list_r.append(i)
        if lane_o == 4:
            for i in self.actor_list:
                if abs(i.get_transform().location.y - 0.00618 * i.get_transform().location.x - 18.23) < 1.8:
                    list_o.append(i)
                if abs(i.get_transform().location.y - 0.00618 * i.get_transform().location.x - 14.63) < 1.8:
                    list_r.append(i)
        """
        Ranking the lanes
        """
        if len(list_o) > 1:
            for j in range(len(list_o)):
                for k in range(0, len(list_o) - j - 1):
                    if list_o[k].get_transform().location.x < list_o[k + 1].get_transform().location.x:
                        list_o[k], list_o[k + 1] = list_o[k + 1], list_o[k]
        if len(list_r) > 1:
            for j in range(len(list_r)):
                for k in range(0, len(list_r) - j - 1):
                    if list_r[k].get_transform().location.x < list_r[k + 1].get_transform().location.x:
                        list_r[k], list_r[k + 1] = list_r[k + 1], list_r[k]
        if len(list_l) > 1:
            for j in range(len(list_l)):
                for k in range(0, len(list_l) - j - 1):
                    if list_l[k].get_transform().location.x < list_l[k + 1].get_transform().location.x:
                        list_l[k], list_l[k + 1] = list_l[k + 1], list_l[k]
        """
        Find the 5 vehicles from 5 seperate orientation
        """
        for i in list_o:
            if i.get_transform().location.x < actor.get_transform().location.x:
                actor_1 = i
        if len(list_r) > 0:
            for i in range(len(list_r)):
                if list_r[i].get_transform().location.x > actor.get_transform().location.x:
                    actor_r = list_r[i]
                if list_r[i].get_transform().location.x < actor.get_transform().location.x:
                    actor_r_1 = list_r[i]
        else:
            actor_r = 0
            actor_r_1 = 0
        if len(list_l) > 0:
            for i in range(len(list_l)):
                if list_l[i].get_transform().location.x > actor.get_transform().location.x:
                    actor_l = list_l[i]
                if list_l[i].get_transform().location.x < actor.get_transform().location.x:
                    actor_l_1 = list_l[i]
        else:
            actor_l = 0
            actor_l_1 = 0


        return actor_r, actor_l, actor_r_1, actor_l_1, actor_1

    def tick(self):
        while True:
            start = time.time()
            Lock = threading.Lock()
            Lock.acquire()
            for i in range(len(self.actor_list)):
                self.ws.cell(row=self.frame + 1 + self.destroyed_num, column=i + 2).value = - self.actor_list[i].get_velocity().x
                self.ws.cell(row=self.frame + 1 + self.destroyed_num, column=i + 20).value = - self.actor_list[i].get_acceleration().x
            self.frame += 1
            Lock.release()
            end = time.time()
            time.sleep(0.03 - end + start)
            if self.frame >= 800:
               break
        self.wb.save(r"C:\Users\Articial_Idiot\Desktop\data.xlsx")
        #self.destroy_all()

    def observation(self, actor, actor_r, actor_l, actor_r_1, actor_l_1, actor_1):
        stand_still_distance = 5
        h = -1
        ave = 0
        v_ego = -actor.get_velocity().x
        actor_1 = 0
        actor_l_1 = 0
        actor_r_1 = 0
        actor_l = 0
        actor_r = 0
        lane_o = 0
        ay = actor.get_transform().location.y
        ax = actor.get_transform().location.x
        if ay - 0.00618 * ax - 7.43 >= - 1.8 and ay - 0.00618 * ax - 7.43 < 1.8:
            lane_o = 1
        elif ay - 0.00618 * ax - 11.03 >= -1.8 and ay - 0.00618 * ax - 11.03 < 1.8:
            lane_o = 2
        elif ay - 0.00618 * ax - 14.63 >= -1.8 and ay - 0.00618 * ax - 14.63 < 1.8:
            lane_o = 3
        elif ay - 0.00618 * ax - 18.23 >= -1.8 and ay - 0.00618 * ax - 18.23 < 1.8:
            lane_o = 4

        """
        observation of 5 headertimes 
        """
        if actor_1 != 0:
            h_ = (-actor_1.get_transform().location.x + actor.get_transform().location.x - stand_still_distance) / v_ego
        else:
            h_ = h
        if actor_l_1 != 0:
            h_l_1 = (-actor_l_1.get_transform().location.x + actor.get_transform().location.x - stand_still_distance) / v_ego
        else:
            h_l_1 = h
        if actor_l != 0:
            h_l = (-actor.get_transform().location.x + actor_l.get_transform().location.x - stand_still_distance) / (-actor_l.get_velocity().x)
        else:
            h_l = h
        if actor_r_1 != 0:
            h_r_1 = (-actor_r_1.get_transform().location.x + actor.get_transform().location.x - stand_still_distance) / v_ego
        else:
            h_r_1 = h
        if actor_r != 0:
            h_r = (-actor.get_transform().location.x + actor_r.get_transform().location.x - stand_still_distance) / (-actor_r.get_velocity().x)
        else:
            h_r = h
        """
        discretelization
        """
        if h_ == h:
            x_ = 1
        elif h_ > 0 and h_ < 0.3:
            x_ = 2
        elif h_ >= 0.3 and h_ < 8:
            x_ = 3
        else:
            x_ = 4

        if h_r == h:
            x_r = 1
        elif h_r > 0 and h_r < 0.3:
            x_r = 2
        elif h_r >= 0.3 and h_r < 8:
            x_r = 3
        else:
            x_r = 4

        if h_r_1 == h:
            x_r_1 = 1
        elif h_r_1 > 0 and h_r_1 < 0.3:
            x_r_1 = 2
        elif h_r_1 >= 0.3 and h_r_1 < 8:
            x_r_1 = 3
        else:
            x_r_1 = 4

        if h_l == h:
            x_l = 1
        elif h_l > 0 and h_l < 0.3:
            x_l = 2
        elif h_l >= 0.3 and h_l < 8:
            x_l = 3
        else:
            x_l = 4

        if h_l_1 == h:
            x_l_1 = 1
        elif h_l_1 > 0 and h_l_1 < 0.3:
            x_l_1 = 2
        elif h_l_1 >= 0.3 and h_l_1 < 8:
            x_l_1 = 3
        else:
            x_l_1 = 4
        xt = [x_, x_r, x_r_1, x_l, x_l_1]
        s = sum([x * (4 ** i) for i, x in enumerate(xt)])
        return s

    def reward(self, call):
        ave = 0
        if self.call == 1:
            try:
                # Lock = threading.Lock()
                # Lock.acquire()
                for i in range(len(self.actor_list)):
                    try:
                        count = 0
                        if self.actor_list[i].get_transform().location.x < 235 and self.actor_list[i].get_transform().location.x > -375:
                            count += 1
                            ave = (ave * count - self.actor_list[i].get_velocity().x) / (count + 1)
                    except:
                        pass
                reward = ave - self.v_init - self.crash * 1000
                # Lock.release()
            except:
                reward = 0
                print("wrong when getting reward")
        else:
            reward = 0

        return reward

    def get_action(self, state, action, reward, next_state):
        epsilon = 0.5 * (0.9 ** self.episode)  #epsilon-greedy
        if epsilon <= np.random.uniform(0, 1):
            next_action = np.argmax(self.q_table[next_state])
        else:
            next_action = np.random.choice([0, 1, 2, 3])

        alpha = 0.2  # learning rate
        gamma = 0.9  # decaying parameter
        self.q_table[state, action] = (1 - alpha) * self.q_table[state, action] + alpha * (
                reward + gamma * self.q_table[next_state, next_action])
        # -------------------------------------------------------------------------------------------
        return next_action

    def get_v_init(self):
        ave = 0
        s = 0
        try:
            Lock = threading.Lock()
            Lock.acquire()
            for i in self.actor_list:
                s = s + (-i.get_velocity().x)
            ave = s/len(self.actor_list)
            self.v_init = ave
            Lock.release()
        except:
            print("error when getting initial speed")

    def CACC_controller(self, actor, lane_d):
        ebx = []
        eba = []
        eby = []
        a_1__list_a = []
        T = 0
        B = 0
        steering = 0
        qi_buffer = []
        while True:
            #If applying optimization
            if self.call == 1:
                try:
                    actor_1 = 0
                    actor_r = 0
                    actor_l = 0
                    actor_r_1 = 0
                    actor_l_1 = 0
                    actor_r, actor_l, actor_r_1, actor_l_1, actor_1 = self.lane_ob(actor)
                    state = self.observation(actor, actor_r, actor_l, actor_r_1, actor_l_1, actor_1)
                    action = np.argmax(self.q_table[state])
                    episode_reward = 0
                    while True:
                        for n in range(33*7):
                            try:
                                start = 0
                                end = 1
                                try:
                                    start = time.time()
                                    self._flag = 1
                                    self.cond.acquire()
                                    #start = time.time()
                                    actor_1 = 0
                                    actor_r = 0
                                    actor_l = 0
                                    actor_r_1 = 0
                                    actor_l_1 = 0

                                    ax = actor.get_transform().location.x
                                    ay = actor.get_transform().location.y
                                    actor_r, actor_l, actor_r_1, actor_l_1, actor_1 = self.lane_ob(actor)
                                except:
                                    break

                                # This is how it tracks
                                if action == 0:
                                    if actor_1 != 0:
                                        h__ = (-actor_1.get_transform().location.x + ax - 5) / (-actor.get_velocity().x+0.1)
                                        if h__ < 8:
                                            T, B, eba, ebx, a_1__list_a, qi_buffer = self.x_controller(actor, actor_1, eba, ebx, a_1__list_a, qi_buffer)
                                            steering, eby = self.y_controller(actor, lane_d, eby)
                                        else:
                                            T = 0.9
                                            B = 0
                                            steering, eby = self.y_controller(actor, lane_d, eby)
                                    else:
                                        T = random.uniform(0.7, 0.9)
                                        B = 0
                                        steering, eby = self.y_controller(actor, lane_d, eby)

                                # change to left lane
                                elif action == 1:
                                    if actor_1 != 0 and actor_l_1 == 0:
                                        h__ = (-actor_1.get_transform().location.x + ax - 5) / (-actor.get_velocity().x + 0.1)
                                        if h__ < 8:
                                            T, B, eba, ebx, a_1__list_a, qi_buffer = self.x_controller(actor, actor_1, eba, ebx, a_1__list_a, qi_buffer)
                                            steering, eby = self.y_controller(actor, lane_d + 1, eby)
                                        else:
                                            T = 0.9
                                            B = 0
                                            steering, eby = self.y_controller(actor, lane_d + 1, eby)
                                    elif actor_1 == 0 and actor_l_1 != 0:
                                        h__ = (-actor_l_1.get_transform().location.x + ax - 5) / (-actor.get_velocity().x + 0.1)
                                        if h__ < 8:
                                            T, B, eba, ebx, a_1__list_a, qi_buffer = self.x_controller(actor, actor_l_1, eba, ebx, a_1__list_a, qi_buffer)
                                            steering, eby = self.y_controller(actor, lane_d + 1, eby)
                                        else:
                                            T = 0.9
                                            B = 0
                                            steering, eby = self.y_controller(actor, lane_d + 1, eby)
                                    elif actor_1 != 0 and actor_l_1 != 0:
                                        h__1 = (-actor_1.get_transform().location.x + ax - 5) / (-actor.get_velocity().x + 0.1)
                                        h__2 = (-actor_l_1.get_transform().location.x + ax - 5) / (-actor.get_velocity().x + 0.1)
                                        if h__1 < h__2:
                                            h__ = h__1
                                            target = actor_1
                                        else:
                                            h__ = h__2
                                            target = actor_l_1
                                        if h__ < 8:
                                            T, B, eba, ebx, a_1__list_a, qi_buffer = self.x_controller(actor, target, eba, ebx, a_1__list_a, qi_buffer)
                                            steering, eby = self.y_controller(actor, lane_d + 1, eby)
                                        else:
                                            T = 0.9
                                            B = 0
                                            steering, eby = self.y_controller(actor, lane_d + 1, eby)
                                    else:
                                        if lane_d != 4:
                                            T = random.uniform(0.7, 0.9)
                                            B = 0
                                            steering, eby = self.y_controller(actor, lane_d + 1, eby)
                                        elif lane_d == 4:
                                            T = random.uniform(0.7, 0.9)
                                            B = 0
                                            steering, eby = self.y_controller(actor, lane_d, eby)
                                            # print("left change failure")

                                # change to right lane
                                elif action == 2:
                                    if actor_1 != 0 and actor_r_1 == 0:
                                        h__ = (-actor_1.get_transform().location.x + ax - 5) / (-actor.get_velocity().x + 0.1)
                                        if h__ < 8:
                                            T, B, eba, ebx, a_1__list_a, qi_buffer = self.x_controller(actor, actor_1, eba, ebx, a_1__list_a, qi_buffer)
                                            steering, eby = self.y_controller(actor, lane_d - 1, eby)
                                        else:
                                            T = 0.9
                                            B = 0
                                            steering, eby = self.y_controller(actor, lane_d - 1, eby)
                                    elif actor_1 == 0 and actor_r_1 != 0:
                                        h__ = (-actor_r_1.get_transform().location.x + ax - 5) / (-actor.get_velocity().x + 0.1)
                                        if h__ < 8:
                                            T, B, eba, ebx, a_1__list_a, qi_buffer = self.x_controller(actor, actor_r_1, eba, ebx, a_1__list_a, qi_buffer)
                                            steering, eby = self.y_controller(actor, lane_d - 1, eby)
                                        else:
                                            T = 0.9
                                            B = 0
                                            steering, eby = self.y_controller(actor, lane_d - 1, eby)
                                    elif actor_1 != 0 and actor_r_1 != 0:
                                        h__1 = (-actor_1.get_transform().location.x + ax - 5) / (-actor.get_velocity().x + 0.1)
                                        h__2 = (-actor_r_1.get_transform().location.x + ax - 5) / (-actor.get_velocity().x + 0.1)
                                        if h__1 < h__2:
                                            h__ = h__1
                                            target = actor_1
                                        else:
                                            h__ = h__2
                                            target = actor_r_1
                                        if h__ < 8:
                                            T, B, eba, ebx, a_1__list_a, qi_buffer = self.x_controller(actor, target, eba, ebx, a_1__list_a, qi_buffer)
                                            steering, eby = self.y_controller(actor, lane_d - 1, eby)
                                        else:
                                            T = 0.9
                                            B = 0
                                            steering, eby = self.y_controller(actor, lane_d - 1, eby)
                                    else:
                                        if lane_d != 4:
                                            T = random.uniform(0.7, 0.9)
                                            B = 0
                                            steering, eby = self.y_controller(actor, lane_d - 1, eby)
                                        elif lane_d == 4:
                                            T = random.uniform(0.7, 0.9)
                                            B = 0
                                            steering, eby = self.y_controller(actor, lane_d, eby)
                                            # print("右变道失败")

                                # If deceleration
                                elif action == 3:
                                    T = 0
                                    B = 0
                                    sterring, eby = self.y_controller(actor, lane_d, eby)


                                actor.apply_control(carla.VehicleControl(throttle=T, brake=B, steer=steering))
                                # self.tick()
                                self._flag = 0
                                end = time.time()
                                if end - start < 0.03:
                                    self.cond.wait(0.03 - end + start)
                            except:
                                try:
                                    T = random.uniform(0.7, 0.9)
                                    B = 0
                                    actor.apply_control(carla.VehicleControl(throttle=T, brake=B, steer=steering))
                                except:
                                    pass
                            finally:
                                self._flag = 0
                                self.cond.release()
                        next_state = self.observation(actor, actor_r, actor_l, actor_r_1, actor_l_1, actor_1)
                        reward = self.reward(self.call)
                        next_action = self.get_action(state, action, reward, next_state)
                        action = next_action
                        state = next_state
                except:
                    break




            # If optimization is off
            else:
                try:
                    for n in range(33*7):
                        start = 0
                        end = 1
                        try:
                            start = time.time()
                            self._flag = 1
                            self.cond.acquire()
                            # start = time.time()
                            actor_1 = 0
                            actor_r = 0
                            actor_l = 0
                            actor_r_1 = 0
                            actor_l_1 = 0

                            ax = actor.get_transform().location.x
                            ay = actor.get_transform().location.y
                            actor_r, actor_l, actor_r_1, actor_l_1, actor_1 = self.lane_ob(actor)

                            if actor_1 != 0:
                                h__ = (-actor_1.get_transform().location.x + ax - 5) / (-actor.get_velocity().x + 0.1)
                                if h__ < 8:
                                    T, B, eba, ebx, a_1__list_a, qi_buffer = self.x_controller(actor, actor_1, eba, ebx,
                                                                                               a_1__list_a, qi_buffer)
                                    steering, eby = self.y_controller(actor, lane_d, eby)
                                else:
                                    T = 0.9
                                    B = 0
                                    steering, eby = self.y_controller(actor, lane_d, eby)
                            else:
                                T = random.uniform(0.7, 0.9)
                                B = 0
                                steering, eby = self.y_controller(actor, lane_d, eby)
                            actor.apply_control(carla.VehicleControl(throttle=T, brake=B, steer=steering))
                            # self.tick()
                            self._flag = 0
                            end = time.time()
                            if end - start < 0.03:
                                self.cond.wait(0.03 - end + start)
                        except:
                            try:
                                T = random.uniform(0.7, 0.9)
                                B = 0
                                actor.apply_control(carla.VehicleControl(throttle=T, brake=B, steer=steering))
                            except:
                                pass
                        finally:
                            self._flag = 0
                            self.cond.release()
                except:
                    break

    def destruction(self):

        while True:
            # start = time.time()
            self.cond.acquire()
            if self._flag == 0:
                start = time.time()
                if len(self.actor_list) != 0:
                    for actor in self.actor_list:
                        if abs(self.rot(actor.get_transform().rotation.yaw)) >= np.pi/9 or actor.get_transform().location.x < - 400:
                            Lock = threading.Lock()
                            Lock.acquire()
                            for i in range(len(self.actor_list)):
                                if abs(self.actor_list[i].get_transform().location.x - actor.get_transform().location.x) < 1 and abs(
                                        self.actor_list[i].get_transform().location.y - actor.get_transform().location.y) < 0.3:
                                    self.actor_list.pop(i)
                                    break
                            self.destroyed_num += 1
                            actor.destroy()
                end = time.time()
                self.cond.wait(0.03 - end + start)
            else:
                self.cond.wait(0.01)
            self.cond.release()


    def run(self):
        a = 0
        for self.episode in range(500):
            self.reward_sum = 0
            self.client = carla.Client("localhost", 2000)
            self.client.set_timeout(30.0)
            # get world
            self.world = self.client.load_world('Town01')
            self.world = self.client.load_world('Town04')

            # settings = self.world.get_settings()
            # settings.no_rendering_mode = True
            # self.world.apply_settings(settings)
            print("load complete")

            thread2 = threading.Thread(target=self.destruction)
            thread2.start()

            thread3 = threading.Thread(target = self.crash_detection)
            thread3.start()

            print("episode = " + str(self.episode))
            self.reward_sum = 0
            self.call = 0
            self.v_init = 0
            self.destroyed_num = 0
            self.shock = 0
            self.frame = 0
            self._flag = 0
            self.crash = 0

            start = time.time()
            a += 1
            while True:
                for j in range(1, 5):
                    time.sleep(1)
                    i = random.randint(1,4)
                    #self.note.write(str(self.episode) + "\t" + str(self.reward_sum) + "\t" + str(self.crash) + "\n")
                    try:
                        self.spawn_vehicle(i)
                        thread1 = threading.Thread(target=self.CACC_controller, args=(self.actor_list[-1], i))
                        thread1.start()
                    except:
                        pass
                    finally:
                        self.reward_sum += self.reward(self.call)
                        end = time.time()
                        print(str(end - start) + '\t' + str(self.reward_sum) + "\t" + str(self.crash))
                if end - start > 40 and self.call == 0:
                    Lock = threading.Lock()
                    Lock.acquire()
                    self.call = 1
                    self.get_v_init()
                    Lock.release()
                if self.crash == 1 or end - start > 70:
                    break

            #self.note.write(str(self.episode) + "\t" + str(self.reward_sum) + "\t" + str(self.crash) + "\n")
            self.note.cell(row=a, column=1).value = self.episode
            self.note.cell(row=a, column=2).value = self.reward_sum
            self.wb.save(r"C:\Users\Articial_Idiot\Desktop\data.xlsx")






    def destroy_all(self):
        Lock = threading.Lock()
        Lock.acquire()
        for i in range(len(self.actor_list)):
            self.actor_list[i].destroy()
            self.actor_list.pop(i)
        Lock.release()





"""
main function
"""
thread1 = CACC(note= wb)
thread1.start()
